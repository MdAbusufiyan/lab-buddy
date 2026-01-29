import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from wsgiref import headers
import requests
from PIL import Image, ImageTk, ImageDraw
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
import os
import sys
import json
import threading
import time
import webbrowser
import re
from openpyxl.utils import get_column_letter
import hashlib


SEARCH_PLACEHOLDER = "Use me for search…"
PLACEHOLDER_COLOR = "gray"
NORMAL_COLOR = "black"
CACHE_FILE = "chemical_cache.json"
CACHE_SIG_FILE = "chemical_cache.sig"

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

class PubChemScraperApp:
    def __init__(self, root):
        self.root = root
        self.root.title("LAB Buddy")
        self.root.geometry("1050x750")
        self.root.minsize(1350, 750)
        self.root.state("zoomed")
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        try:
            self.root.iconbitmap(resource_path("ico.ico"))
        except Exception:
            pass
        self.excel_file = None
        self.current_data = None
        self.excel_frame_visible = False
        self.suggestion_confirmed = False
        self.suggestion_popup = None
        self.search_in_progress = False
        self.header_bg_image = None
        try:
            img = Image.open(resource_path("header_polymer.png"))
            img = img.resize((1600, 70), Image.Resampling.LANCZOS)
            self.header_bg_image = ImageTk.PhotoImage(img)
        except Exception as e:
            pass
        self.suggestions = []
        self.suggestion_listbox = None
        self.autocomplete_active = False
        self.last_search_time = 0
        self.last_searched_query = None

        self.include_cas = tk.BooleanVar(value=True)
        self.include_formula = tk.BooleanVar(value=True)
        self.include_iupac = tk.BooleanVar(value=False)
        self.include_smiles = tk.BooleanVar(value=False)
        self.include_molweight = tk.BooleanVar(value=True)
        self.include_density = tk.BooleanVar(value=True)
        self.include_quantity = tk.BooleanVar(value=True)
        self.include_equivalence = tk.BooleanVar(value=True)
        self.include_image_link = tk.BooleanVar(value=False)

        self.title_var = tk.StringVar()
        self.formula_var = tk.StringVar()
        self.cas_var = tk.StringVar()
        self.molweight_var = tk.StringVar()
        self.density_var = tk.StringVar()

        self.create_widgets()

        self.cache = {}

        try:
            with open(CACHE_FILE, "rb") as f:
                raw = f.read()

            with open(CACHE_SIG_FILE, "r") as sig:
                stored_hash = sig.read().strip()

            if self.compute_hash(raw) != stored_hash:
                raise ValueError("Cache integrity check failed")

            self.cache = json.loads(raw.decode("utf-8"))
            self.log("✓ Cache loaded (integrity verified)")

        except Exception:
            self.cache = {}
            self.log("⚠ Cache invalid or missing — ignored safely")

        # ✅ ALWAYS build indices (outside except)
        self.cas_index = {}
        self.iupac_index = {}
        self.smiles_index = {}

        for key, data in self.cache.items():
            if data.get("cas"):
                self.cas_index[data["cas"].lower()] = key

            if data.get("iupac"):
                self.iupac_index[self.normalize_key(data["iupac"])] = key

            if data.get("smiles"):
                self.smiles_index[data["smiles"]] = key

    def on_close(self):
        if messagebox.askyesno(
            "Exit LAB Buddy",
            "Any unsaved data will be lost.\n\nDo you want to exit LAB Buddy?"
        ):
            self.root.destroy()

    def open_dev_profile(self, event=None):
        webbrowser.open_new(
            "https://www.linkedin.com/in/sufiyanabu/"
        )
    
    def make_circular_image(self, img, size=180, border=6):
        img = img.resize((size, size), Image.Resampling.LANCZOS).convert("RGBA")

        # Create circular mask
        mask = Image.new("L", (size, size), 0)
        draw = ImageDraw.Draw(mask)
        draw.ellipse((0, 0, size, size), fill=255)
        img.putalpha(mask)

        # Create black background for border
        final_size = size + border * 2
        background = Image.new("RGBA", (final_size, final_size), (0, 0, 0, 255))
        background.paste(img, (border, border), img)

        return background

    def cache_suggestions(self, query, limit=6):
        q = self.normalize_key(query)
        results = []

        for key, data in self.cache.items():
            name = self.normalize_key(data.get("name", ""))
            if name.startswith(q):
                results.append(data["name"])
                if len(results) >= limit:
                    break

        return results
    

    def create_widgets(self):

        # ================= HEADER BAR =================
        header_frame = tk.Frame(self.root, height=70)
        header_frame.grid(row=0, column=0, sticky="ew")
        header_frame.grid_propagate(False)

        # Header background image
        if self.header_bg_image:
            bg = tk.Label(header_frame, image=self.header_bg_image)
            bg.place(relx=0, rely=0, relwidth=1, relheight=1)

        # App title
        tk.Label(
            header_frame,
            text="LAB Buddy",
            font=("Segoe UI", 16, "bold"),
            fg="white",
            bg="#3D91AD"
        ).pack(side="top", pady=(8, 0))

        about_btn = tk.Button(
            header_frame,
            text="About",
            font=("Segoe UI", 10, "bold"),
            bg="#E6E6E6",
            fg="#000000",
            relief="raised",
            bd=2,
            highlightthickness=0,
            activebackground="#DADADA",
            activeforeground="#000000",
            cursor="hand2",
            command=self.open_about_window
        )

        about_btn.grid(
            row=0,
            column=0,
            sticky="w",
            padx=10,
            pady=10
        )
        about_btn.bind("<Enter>", lambda e: about_btn.config(bg="#DADADA"))
        about_btn.bind("<Leave>", lambda e: about_btn.config(bg="#E6E6E6"))

        help_btn = tk.Button(
            header_frame,
            text="Help",
            font=("Segoe UI", 10, "bold"),
            bg="#CED2D6",
            fg="#000000",
            relief="raised",
            bd=2,
            highlightthickness=0,
            activebackground="#DADADA",
            activeforeground="#000000",
            cursor="hand2",
            command=self.open_help_pdf
        )

        help_btn.grid(
            row=0,
            column=1,
            sticky="e",
            padx=10,
            pady=10
        )

        header_frame.grid_columnconfigure(1, weight=1)

        main_frame = tk.Frame(self.root)
        main_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=10)

        self.root.grid_rowconfigure(1, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        main_frame.grid_columnconfigure(0, weight=5)  # LEFT = wider
        main_frame.grid_columnconfigure(1, weight=1, minsize=300)
        main_frame.grid_rowconfigure(0, weight=1)

        left_container = tk.Frame(main_frame)
        left_container.grid(row=0, column=0, sticky="nsew", padx=(0, 10))

        self.left_canvas = tk.Canvas(left_container, highlightthickness=0)
        self.left_canvas.pack(side="left", fill="both", expand=True)

        left_scrollbar = tk.Scrollbar(left_container, orient="vertical", command=self.left_canvas.yview)
        left_scrollbar.pack(side="right", fill="y")

        self.left_canvas.configure(yscrollcommand=left_scrollbar.set)

        left_frame = tk.Frame(self.left_canvas)
        self.left_canvas.create_window((0, 0), window=left_frame, anchor="nw")
        self.left_canvas.bind(
            "<Configure>",
            lambda e: self.left_canvas.itemconfig("all", width=e.width)
        )

        # --- SEARCH HIGHLIGHT BOX ---
        search_container = tk.Frame(
            left_frame,
            bd=1,
            relief="groove",
            padx=4,
            pady=6
        )
        search_container.pack(fill="x", pady=(2, 10))

        tk.Label(search_container, text="Search:", font=("Arial", 11, "bold")).pack(anchor="w", pady=(0, 2))

        self.name_frame = tk.Frame(search_container)
        self.name_frame.pack(fill="x")

        # --- Search row split into 2 columns ---
        search_row = tk.Frame(search_container)
        search_row.pack(fill="x")

        # LEFT: Search box
        search_box = tk.Frame(search_row)
        search_box.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        search_row.grid_columnconfigure(0, weight=3)

        self.name_entry = tk.Entry(
            search_box,
            font=("Arial", 11),
            fg=PLACEHOLDER_COLOR
        )
        self.name_entry.pack(side="left", fill="x", expand=True)

        self.name_entry.insert(0, SEARCH_PLACEHOLDER)

        search_btn = tk.Button(
            search_box,
            text="🔍",
            command=self.search_chemical,
            font=("Arial", 11),
            width=3
        )
        search_btn.pack(side="right", padx=(5, 0))

        ttk.Separator(left_frame, orient="horizontal").pack(fill="x", pady=10)

         # --- INFO ROW 1: TITLE + CAS ---
        info_row1 = tk.Frame(left_frame)
        info_row1.pack(fill="x", pady=(0, 8))

        info_row2 = tk.Frame(left_frame)
        info_row2.pack(fill="x", pady=(0, 10))

        cas_frame = tk.Frame(info_row1)
        cas_frame.pack(side="left", fill="x", expand=True, padx=(3, 3))

        tk.Label(
            cas_frame,
            text="CAS No:",
            font=("Arial", 9, "bold")
        ).pack(anchor="w")

        cas_row = tk.Frame(cas_frame)
        cas_row.pack(fill="x")

        self.cas_entry = tk.Entry(
            cas_row,
            textvariable=self.cas_var,
            font=("Arial", 10),
            state="readonly",
            relief="sunken"
        )
        self.cas_entry.pack(side="left", fill="x", expand=True)

        tk.Button(
            cas_row,
            text="⧉",
            width=2,
            command=lambda: self.copy_to_clipboard(self.cas_entry)
        ).pack(side="right", padx=3)

        title_box = tk.Frame(info_row1)
        title_box.pack(side="left", fill="x", expand=True, padx=(3, 3))

        tk.Label(
            title_box,
            text="Title:",
            font=("Arial", 9, "bold")
        ).pack(anchor="w")

        title_row = tk.Frame(title_box)
        title_row.pack(fill="x")

        self.title_entry = tk.Entry(
            title_row,
            textvariable=self.title_var,
            font=("Arial", 10),
            state="readonly",
            relief="sunken"
        )
        self.title_entry.pack(side="left", fill="x", expand=True)

        tk.Button(
            title_row,
            text="⧉",
            width=2,
            command=lambda: self.copy_to_clipboard(self.title_entry) 
            ).pack(side="right", padx=3)

        readonly_bg = self.title_entry.cget("bg")
        self.name_entry.bind('<KeyRelease>', self.on_key_release)
        self.name_entry.bind('<Return>', self.on_enter_pressed)
        self.name_entry.bind('<Down>', self.on_down_key)
        self.name_entry.bind('<Escape>', self.hide_suggestions)
        self.name_entry.bind("<FocusIn>", self._clear_search_placeholder)
        self.name_entry.bind("<FocusOut>", self._restore_search_placeholder)

        formula_frame = tk.Frame(info_row2)
        formula_frame.pack(side="left", fill="x", expand=True, padx=(0, 3))

        molweight_frame = tk.Frame(info_row2)
        molweight_frame.pack(side="left", fill="x", expand=True, padx=(3, 0))

        tk.Label(
            molweight_frame,
            text="Mol. Weight:",
            font=("Arial", 9, "bold")
        ).pack(anchor="w")

        mw_row = tk.Frame(molweight_frame)
        mw_row.pack(fill="x")

        self.molweight_entry = tk.Entry(
            mw_row,
            textvariable=self.molweight_var,
            font=("Arial", 10),
            state="readonly",
            relief="sunken"
        )
        self.molweight_entry.pack(side="left", fill="x", expand=True)

        tk.Button(
            mw_row,
            text="⧉",
            width=2,
            command=lambda: self.copy_to_clipboard(self.molweight_entry)
        ).pack(side="right", padx=3)

        density_frame = tk.Frame(info_row2)
        density_frame.pack(side="left", fill="x", expand=True, padx=(3, 0))

        tk.Label(
            density_frame,
            text="Density:",
            font=("Arial", 9, "bold")
        ).pack(anchor="w")

        density_row = tk.Frame(density_frame)
        density_row.pack(fill="x")

        self.density_entry = tk.Entry(
            density_row,
            textvariable=self.density_var,
            font=("Arial", 10),
            state="readonly",
            relief="sunken"
        )
        self.density_entry.pack(side="left", fill="x", expand=True)

        tk.Button(
            density_row,
            text="⧉",
            width=2,
            command=lambda: self.copy_to_clipboard(self.density_entry)
        ).pack(side="right", padx=3)

        tk.Label(
            formula_frame,
            text="Molecular Formula:",
            font=("Arial", 9, "bold")
        ).pack(anchor="w")

        formula_row = tk.Frame(formula_frame)
        formula_row.pack(fill="x")

        self.formula_entry = tk.Entry(
            formula_row,
            textvariable=self.formula_var,
            font=("Arial", 10),
            state="readonly",
            relief="sunken"
        )
        self.formula_entry.pack(side="left", fill="x", expand=True)

        tk.Button(
            formula_row,
            text="⧉",
            width=2,
            command=lambda: self.copy_to_clipboard(self.formula_entry)
        ).pack(side="right", padx=3)

        tk.Label(left_frame, text="IUPAC Name:", font=("Arial", 10, "bold")).pack(anchor="w", pady=(5, 0))

        iupac_text_frame = tk.Frame(left_frame, relief="sunken", bd=2,  bg=readonly_bg)
        iupac_text_frame.pack(fill="x", pady=(0, 10))

        iupac_row = tk.Frame(iupac_text_frame, bg=readonly_bg)
        iupac_row.pack(fill="x")

        self.iupac_text = tk.Text(
            iupac_row,
            height=2,
            wrap=tk.NONE,
            font=("Arial", 9),
            bg=readonly_bg,
            relief="flat",
            bd=0,
            highlightthickness=0,
            highlightbackground=readonly_bg
        )
        self.iupac_text.pack(side="left", fill="both", expand=True)
        self.iupac_text.config(insertbackground=readonly_bg)

        tk.Button(
            iupac_row,
            text="⧉",
            width=2,
            command=lambda: self.copy_to_clipboard(self.iupac_text)
        ).pack(side="right", padx=3)

        tk.Label(left_frame, text="SMILES:", font=("Arial", 10, "bold")).pack(anchor="w", pady=(5, 0))
        smiles_frame = tk.Frame(left_frame, relief="sunken", bd=2, bg=readonly_bg)
        smiles_frame.pack(fill="x", pady=(0, 10))

        smiles_row = tk.Frame(smiles_frame, bg=readonly_bg)
        smiles_row.pack(fill="x")

        self.smiles_text = tk.Text(
            smiles_row,
            height=2,
            wrap=tk.NONE,
            font=("Arial", 9),
            bg=readonly_bg,
            relief="flat",
            bd=0,
            highlightthickness=0,
            highlightbackground=readonly_bg
        )
        self.smiles_text.pack(side="left", fill="both", expand=True)
        self.smiles_text.config(insertbackground=readonly_bg)

        tk.Button(
            smiles_row,
            text="⧉",
            width=2,
            command=lambda: self.copy_to_clipboard(self.smiles_text)
        ).pack(side="right", padx=3)

        tk.Label(left_frame, text="Primary Hazards (GHS Pictograms):", font=("Arial", 11, "bold")).pack(anchor="w", pady=(5, 0))

        self.hazard_frame = tk.Frame(left_frame, bg="black", relief="sunken", bd=2, height=140)
        self.hazard_frame.pack(fill="x", pady=(0, 10))
        self.hazard_frame.pack_propagate(False)

        self.hazard_label = tk.Label(self.hazard_frame, text="No hazard data", bg="black", fg="gray")
        self.hazard_label.pack(expand=True)

        tk.Label(left_frame, text="GHS Hazard Statements:", font=("Arial", 10, "bold")).pack(anchor="w")
        hazard_text_frame = tk.Frame(left_frame, bg="white", relief="sunken", bd=2)
        hazard_text_frame.pack(fill="both", expand=True, pady=(0, 10))

        hazard_scroll = tk.Scrollbar(hazard_text_frame)
        hazard_scroll.pack(side="right", fill="y")

        self.hazard_text = tk.Text(hazard_text_frame, height=6, yscrollcommand=hazard_scroll.set,
                                   font=("Arial", 9), wrap=tk.WORD)
        self.hazard_text.pack(fill="both", expand=True)
        hazard_scroll.config(command=self.hazard_text.yview)

        # --- Excel controls (hidden by default) ---
        self.excel_frame = tk.LabelFrame(left_frame, text="Excel File", padx=10, pady=10)

        self.file_label = tk.Label(self.excel_frame, text="No file selected", fg="gray")
        self.file_label.pack(side="left", padx=5)

        create_btn = tk.Button(
            self.excel_frame,
            text="Create New Excel",
            command=self.prompt_column_selection,
            bg="#27AE60",
            fg="white",
            padx=10,
            pady=5
        )
        create_btn.pack(side="right", padx=5)

        load_btn = tk.Button(
            self.excel_frame,
            text="Load Existing Excel",
            command=self.load_excel_file,
            bg="#3498DB",
            fg="white",
            padx=10,
            pady=5
        )
        load_btn.pack(side="right", padx=5)

        self.excel_frame.pack_forget()

        button_frame = tk.Frame(left_frame)
        button_frame.pack(fill="x", pady=10)

        save_btn = tk.Button(button_frame, text="Save", command=self.add_to_excel,
                             bg="#8E44AD", fg="white", font=("Arial", 11, "bold"),
                             height=2)
        save_btn.pack(side="left", fill="x", expand=True)

        excel_toggle_btn = tk.Button(
            button_frame,
            text="Open Log File ▼",
            command=self.toggle_excel_frame,
            bg="#34495E",
            fg="white",
            font=("Arial", 11, "bold"),
            height=2
        )
        excel_toggle_btn.pack(side="left", fill="x", expand=True, padx=(5, 0))

        more_btn = tk.Button(
            button_frame,
            text="MORE",
            command=self.open_pubchem_page,
            bg="#95A5A6",
            fg="white",
            font=("Arial", 11, "bold"),
            height=2
        )
        more_btn.pack(side="left", fill="x", expand=True, padx=(5, 0))

        # ---------- DATA SOURCE (SUBTLE) ----------
        tk.Label(
            left_frame,
            text="Data sourced from PubChem (NIH).",
            font=("Arial", 9),
            fg="#7F8C8D"   # muted gray
        ).pack(anchor="sw", pady=(2, 8))

        right_frame = tk.Frame(main_frame)
        right_frame.grid(row=0, column=1, sticky="nsew")

        tk.Label(right_frame, text="Structure:", font=("Arial", 11, "bold")).pack(anchor="w", pady=(5, 0))
        self.image_frame = tk.Frame(
            right_frame,
            bg="white",
            relief="sunken",
            bd=2,
            height=450,
            width=400
        )

        self.image_frame.pack(fill="both", pady=(0, 10))
        self.image_frame.pack_propagate(False)

        self.image_label = tk.Label(self.image_frame, text="No image", bg="white", fg="gray")
        self.image_label.pack(expand=True)

        # ---------- IMAGE ACTION BUTTONS (INSIDE IMAGE BOX) ----------
        img_btn_overlay = tk.Frame(self.image_frame, bg="white")
        img_btn_overlay.pack(
            side="bottom",
            anchor="e",
            padx=6,
            pady=6
        )

        tk.Button(
            img_btn_overlay,
            text=" ⧉ Copy Image",
            command=self.copy_image_to_clipboard,
            bg="#3498DB",
            fg="white",
            font=("Arial", 9, "bold"),
            padx=8,
            pady=3
        ).pack(side="left", padx=(0, 5))

        tk.Button(
            img_btn_overlay,
            text=" ↗ Copy Adress",
            command=self.copy_image_to_clipboard_url,
            bg="#5DADE2",
            fg="white",
            font=("Arial", 9, "bold"),
            padx=8,
            pady=3
        ).pack(side="left")


        tk.Label(right_frame, text="Notes:", font=("Arial", 11, "bold")).pack(anchor="w", pady=(5, 0))
        log_frame = tk.Frame(right_frame, relief="sunken", bd=2, height=250)
        log_frame.pack(fill="both", pady=(0 , 10))

        scroll = tk.Scrollbar(log_frame)
        scroll.pack(side="right", fill="y")

        self.log_text = tk.Text(log_frame, yscrollcommand=scroll.set,
                               font=("Courier", 9), wrap=tk.WORD)
        self.log_text.pack(fill="both", expand=True)
        self.log_text.lift()
        scroll.config(command=self.log_text.yview)

        left_frame.bind(
            "<Configure>",
            lambda e: self.left_canvas.configure(scrollregion=self.left_canvas.bbox("all"))
        )

        self.left_canvas.bind_all(
            "<MouseWheel>",
            lambda e: self.left_canvas.yview_scroll(-1 * int(e.delta / 120), "units")
        )

    def _clear_search_placeholder(self, event):
        if self.name_entry.get() == SEARCH_PLACEHOLDER:
            self.name_entry.delete(0, tk.END)
            self.name_entry.config(fg=NORMAL_COLOR)

    def _restore_search_placeholder(self, event):
        if not self.name_entry.get().strip():
            self.name_entry.insert(0, SEARCH_PLACEHOLDER)
            self.name_entry.config(fg=PLACEHOLDER_COLOR)

    def open_about_window(self):
        win = tk.Toplevel(self.root)
        win.title("About LAB Buddy")
        win.geometry("720x220")

        win.update_idletasks()

        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (win.winfo_width() // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (win.winfo_height() // 2)

        win.geometry(f"+{x}+{y}")

        win.resizable(False, False)
        win.transient(self.root)
        win.grab_set()

        try:
            win.iconbitmap(resource_path("ico.ico"))
        except Exception:
            pass

        # Main container
        container = tk.Frame(win, bg="#E6E6E6", padx=15, pady=15)
        container.pack(fill="both", expand=True)

        # ================= LEFT SIDE (IMAGE) =================
        left = tk.Frame(container, bg="#E6E6E6")
        left.pack(side="left", padx=(10, 25), pady=10)

        try:
            img = Image.open(resource_path("profile.png"))
            circ_img = self.make_circular_image(img, size=180, border=6)
            photo = ImageTk.PhotoImage(circ_img)

            img_label = tk.Label(
                left,
                image=photo,
                bg="#E6E6E6",
                cursor="hand2"
            )
            img_label.image = photo
            img_label.pack()

            img_label.bind(
                "<Button-1>",
                lambda e: webbrowser.open_new(
                    "https://www.linkedin.com/in/sufiyanabu/"
                )
            )
        except:
            tk.Label(
                left,
                text="Profile",
                fg="white",
                bg="black",
                width=20,
                height=10
            ).pack()

        # ================= RIGHT SIDE (TEXT) =================
        right = tk.Frame(container, bg="#E6E6E6")
        right.pack(side="left", fill="both", expand=True, pady=15)

        def info_row(label, text, url=None, color="#0B5ED7"):
            row = tk.Frame(right, bg="#E6E6E6")
            row.pack(anchor="w", pady=10)

            tk.Label(
                row,
                text=label,
                font=("Segoe UI", 11, "bold"),
                bg="#E6E6E6"
            ).pack(side="left", padx=(0, 10))

            link = tk.Label(
                row,
                text=text,
                fg=color,
                cursor="hand2" if url else "arrow",
                font=("Segoe UI", 10, "underline" if url else "normal"),
                bg="#E6E6E6"
            )
            link.pack(side="left")

            if url:
                link.bind("<Button-1>", lambda e: webbrowser.open_new(url))

        # About Me
        info_row(
            "About Me:",
            "linkedin.com/in/sufiyanabu",
            "https://www.linkedin.com/in/sufiyanabu/",
            color="#7A3E6C"
        )

        # Source
        info_row(
            "Source:",
            "pubchem.ncbi.nlm.nih.gov",
            "https://pubchem.ncbi.nlm.nih.gov"
        )

        # Version + GitHub
        row = tk.Frame(right, bg="#E6E6E6")
        row.pack(anchor="w", pady=10)

        tk.Label(
            row,
            text="Version:",
            font=("Segoe UI", 11, "bold"),
            bg="#E6E6E6"
        ).pack(side="left", padx=(0, 10))

        tk.Label(
            row,
            text="v1.0.0",
            font=("Segoe UI", 10),
            bg="#E6E6E6"
        ).pack(side="left", padx=(0, 15))

        gh = tk.Label(
            row,
            text="github.com/MdAbusufiyan/lab-buddy",
            fg="black",
            cursor="hand2",
            font=("Segoe UI", 10, "underline"),
            bg="#E6E6E6"
        )
        gh.pack(side="left")
        gh.bind(
            "<Button-1>",
            lambda e: webbrowser.open_new(
                "https://github.com/MdAbusufiyan/lab-buddy"
            )
        )

        # Close button (bottom-right)
        close_btn = tk.Button(
            container,
            text="Close",
            width=10,
            command=win.destroy
        )
        close_btn.pack(side="bottom", anchor="e", pady=(10, 0))

    def copy_image_to_clipboard_url(self):
        if not self.current_data:
            messagebox.showwarning(
                "No Image",
                "Please search for a chemical first."
            )
            return

        image_url = self.current_data.get("image")
        if not image_url:
            messagebox.showwarning("No Image", "No image available.")
            return

        self.root.clipboard_clear()
        self.root.clipboard_append(image_url)
        self.log("✓ Image link copied to clipboard")

    def copy_image_to_clipboard(self):
        if not self.current_data:
            messagebox.showwarning(
                "No Image",
                "Please search for a chemical first."
            )
            return

        image_url = self.current_data.get("image")
        if not image_url:
            messagebox.showwarning("No Image", "No image available.")
            return

        try:
            response = requests.get(image_url, timeout=10)
            response.raise_for_status()

            temp_path = os.path.join(
                os.environ.get("TEMP", "."),
                "labbuddy_structure.png"
            )

            with open(temp_path, "wb") as f:
                f.write(response.content)

            os.startfile(temp_path)  # Windows opens image viewer
            self.log("✓ Structure image opened (source: PubChem)")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to open image:\n{e}")
    
    def copy_to_clipboard(self, source):
        if source is None:
            return

        try:
            if isinstance(source, tk.Entry):
                value = source.get()
            elif isinstance(source, tk.Text):
                value = source.get("1.0", tk.END).strip()
            elif isinstance(source, tk.StringVar):
                value = source.get()
            else:
                return

            if not value:
                return

            self.root.clipboard_clear()
            self.root.clipboard_append(value)
            self.log("Copied")

        except Exception:
            pass

    def prompt_column_selection(self):
        win = tk.Toplevel(self.root)
        win.title("Select Excel Columns")

        try:
            win.iconbitmap(resource_path("ico.ico"))
        except:
            pass
        win.title("Select Excel Columns")
        win.geometry("400x400")
        win.resizable(False, False)

        frame = tk.Frame(win)
        frame.pack(pady=10)

        tk.Label(
            frame,
            text="Default columns (always included):",
            font=("Arial", 9, "bold"),
            fg="gray"
        ).pack(anchor="w", pady=(0, 4))

        label = tk.Label(
            frame,
            text="☑ Chemical Name",
            anchor="w"
        )
        label.pack(anchor="w")

        ttk.Separator(frame, orient="horizontal").pack(fill="x", pady=6)

        tk.Checkbutton(frame, text="CAS Number", variable=self.include_cas).pack(anchor="w")
        tk.Checkbutton(frame, text="Molecular Formula", variable=self.include_formula).pack(anchor="w")
        tk.Checkbutton(frame, text="Molecular Weight", variable=self.include_molweight).pack(anchor="w")
        tk.Checkbutton(frame, text="Density", variable=self.include_density).pack(anchor="w")
        tk.Checkbutton(frame, text="Quantity", variable=self.include_quantity).pack(anchor="w")
        tk.Checkbutton(frame, text="Equivalence", variable=self.include_equivalence).pack(anchor="w")
        tk.Checkbutton(frame, text="IUPAC Name", variable=self.include_iupac).pack(anchor="w")
        tk.Checkbutton(frame, text="SMILES", variable=self.include_smiles).pack(anchor="w")
        tk.Checkbutton(frame, text="Image Link", variable=self.include_image_link).pack(anchor="w")


        def confirm():
            win.destroy()
            self.create_excel_file()

        tk.Button(
            win,
            text="Create Excel",
            command=confirm,
            bg="#27AE60",
            fg="white",
            padx=20
        ).pack(pady=10)

    def on_enter_pressed(self, event):
        # First Enter after suggestion = confirm only
        if self.suggestion_confirmed:
            self.suggestion_confirmed = False
            return

        # Otherwise, run search
        threading.Thread(
            target=self.search_chemical,
            daemon=True
        ).start()

    def on_key_release(self, event):
        if event.keysym in ('Return', 'Up', 'Down', 'Left', 'Right', 'Escape', 'Tab'):
            return

        value = self.name_entry.get().strip()
        if len(value) < 2:
            self.hide_suggestions()
            return

        cached = self.cache_suggestions(value)

        if cached:
            self.show_suggestions(cached)
        else:
            threading.Thread(
                target=self.fetch_suggestions,
                args=(value,),
                daemon=True
            ).start()
    
    def open_help_pdf(self):
        try:
            pdf_path = resource_path("LAB_Buddy_Help.pdf")
            os.startfile(pdf_path)  # Windows default PDF viewer
        except Exception:
            messagebox.showerror(
                "Help",
                "Help guide could not be opened."
            )

    def toggle_excel_frame(self):
        if self.excel_frame_visible:
            self.excel_frame.pack_forget()
        else:
            self.excel_frame.pack(fill="x", pady=(5, 0))
            self.root.after(10, lambda: self.left_canvas.yview_moveto(1.0))
        self.excel_frame_visible = not self.excel_frame_visible

    def fetch_suggestions(self, query):
        try:
            url = f"https://pubchem.ncbi.nlm.nih.gov/rest/autocomplete/compound/{query}/json?limit=10"
            response = requests.get(url, timeout=3)

            if response.status_code == 200:
                data = response.json()
                suggestions = []
                if 'dictionary_terms' in data and 'compound' in data['dictionary_terms']:
                    suggestions = data['dictionary_terms']['compound'][:10]

                self.root.after(0, self.show_suggestions, suggestions)
        except:
            pass

    def show_suggestions(self, suggestions):
        if not suggestions or not self.name_entry.get().strip():
            self.hide_suggestions()
            return

        # Destroy old popup if exists
        if self.suggestion_popup:
            self.suggestion_popup.destroy()

        # Create popup
        self.suggestion_popup = tk.Toplevel(self.root)
        self.suggestion_popup.overrideredirect(True)
        self.suggestion_popup.configure(bg="black")

        # Position popup under entry
        x = self.name_entry.winfo_rootx()
        y = self.name_entry.winfo_rooty() + self.name_entry.winfo_height()
        w = self.name_entry.winfo_width()

        self.suggestion_popup.geometry(f"{w}x{min(150, 22*len(suggestions))}+{x}+{y}")

        # Listbox with border + relief
        self.suggestion_listbox = tk.Listbox(
            self.suggestion_popup,
            font=("Arial", 10),
            activestyle="none",
            relief="solid",
            borderwidth=1,
            highlightthickness=0
        )
        self.suggestion_listbox.pack(fill="both", expand=True)

        for s in suggestions:
            self.suggestion_listbox.insert(tk.END, s)

        self.suggestion_listbox.bind("<<ListboxSelect>>", self.on_suggestion_select)
        self.suggestion_listbox.bind("<Escape>", self.hide_suggestions)

        self.autocomplete_active = True

    def on_suggestion_select(self, event):
        if not self.suggestion_listbox:
            return

        selection = self.suggestion_listbox.curselection()
        if selection:
            value = self.suggestion_listbox.get(selection[0])
            self.name_entry.delete(0, tk.END)
            self.name_entry.insert(0, value)
            self.hide_suggestions()
            self.suggestion_confirmed = True
            self.name_entry.focus_set()

    def on_down_key(self, event):
        if self.suggestion_listbox and self.autocomplete_active:
            self.suggestion_listbox.focus()
            self.suggestion_listbox.select_set(0)

    def hide_suggestions(self, event=None):
        if self.suggestion_popup:
            self.suggestion_popup.destroy()
            self.suggestion_popup = None
            self.suggestion_listbox = None
        self.autocomplete_active = False

    def log(self, message):
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update()

    def log_error(self, error_type, error_message, details=""):
        self.log(f"\n{'='*40}")
        self.log(f"❌ ERROR: {error_type}")
        self.log(f"{'='*40}")
        self.log(f"Message: {error_message}")
        if details:
            self.log(f"Details: {details}")
        self.log(f"{'='*40}\n")

    def clear_all(self):
        self.name_entry.delete(0, tk.END)
        self.clear_results()
        self.hide_suggestions()
        self.log(f"\n{'='*40}")
        self.log("✓ All fields cleared")
        self.log(f"{'='*40}\n")
        self.suggestion_confirmed = False

    def clear_results(self):
        # Clear readonly Entry fields
        self.title_var.set("")
        self.formula_var.set("")
        self.cas_var.set("")
        self.molweight_var.set("")
        self.density_var.set("")

        # Clear Text fields safely
        self.set_text_readonly(self.iupac_text, "")
        self.set_text_readonly(self.smiles_text, "")

        self.image_label.config(image="", text="No image")

        for widget in self.hazard_frame.winfo_children():
            widget.destroy()

        self.hazard_label = tk.Label(self.hazard_frame, text="No hazard data", bg="black", fg="gray")
        self.hazard_label.pack(expand=True)

        self.hazard_text.delete(1.0, tk.END)
        self.current_data = None
    
    def fetch_density(self, cid):
        try:
            url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound/{cid}/JSON"
            response = requests.get(url, timeout=15)

            if response.status_code != 200:
                return None, None

            data = response.json()
            sections = data.get("Record", {}).get("Section", [])

            for section in sections:
                if section.get("TOCHeading") == "Chemical and Physical Properties":
                    for sub in section.get("Section", []):
                        if sub.get("TOCHeading") == "Experimental Properties":
                            for prop in sub.get("Section", []):
                                if "density" in prop.get("TOCHeading", "").lower():
                                    for info in prop.get("Information", []):
                                        value = info.get("Value", {})
                                        text = ""

                                        if "StringWithMarkup" in value:
                                            text = value["StringWithMarkup"][0].get("String", "")
                                        elif "StringValue" in value:
                                            text = value["StringValue"]

                                        if not text:
                                            continue

                                        # ---- Extract number ----
                                        match = re.search(r"([\d.]+)", text)
                                        if not match:
                                            continue

                                        density = float(match.group(1))

                                        # ---- Detect temperature ----
                                        temp_c = 25  # default lab temp

                                        if "°f" in text.lower():
                                            temp_f_match = re.search(r"([\d.]+)\s*°\s*f", text.lower())
                                            if temp_f_match:
                                                temp_f = float(temp_f_match.group(1))
                                                temp_c = round((temp_f - 32) * 5 / 9)

                                        elif "°c" in text.lower():
                                            temp_c_match = re.search(r"([\d.]+)\s*°\s*c", text.lower())
                                            if temp_c_match:
                                                temp_c = round(float(temp_c_match.group(1)))

                                        return density, f"g/mL @ {temp_c} °C"


            return None, None

        except Exception as e:
            
            self.log(f"⚠ Density error: {e}")
            return None, None

    def fetch_iupac_name(self, cid):
        iupac_name = "Not available"
        try:
            url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound/{cid}/JSON"
            response = requests.get(url, timeout=15)

            if response.status_code == 200:
                data = response.json()
                sections = data.get('Record', {}).get('Section', [])

                for section in sections:
                    if section.get('TOCHeading') == 'Names and Identifiers':
                        for subsection in section.get('Section', []):
                            if subsection.get('TOCHeading') == 'Computed Descriptors':
                                for info_section in subsection.get('Section', []):
                                    if info_section.get('TOCHeading') == 'IUPAC Name':
                                        for info in info_section.get('Information', []):
                                            value = info.get('Value', {})
                                            if 'StringWithMarkup' in value:
                                                markup_list = value['StringWithMarkup']
                                                if markup_list and len(markup_list) > 0:
                                                    iupac_name = markup_list[0].get('String', 'Not available')
                                                    self.log(f"✓ IUPAC Name found")
                                                    return iupac_name
                                        break
                                break
                        break
        except Exception as e:
            
            self.log(f"⚠ IUPAC: {str(e)}")

        return iupac_name

    def fetch_smiles(self, cid):
        smiles = "Not available"
        try:
            url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound/{cid}/JSON"
            response = requests.get(url, timeout=15)

            if response.status_code == 200:
                data = response.json()
                sections = data.get('Record', {}).get('Section', [])

                for section in sections:
                    if section.get('TOCHeading') == 'Names and Identifiers':
                        for subsection in section.get('Section', []):
                            if subsection.get('TOCHeading') == 'Computed Descriptors':
                                for info_section in subsection.get('Section', []):
                                    if info_section.get('TOCHeading') == 'SMILES':
                                        for info in info_section.get('Information', []):
                                            value = info.get('Value', {})
                                            if 'StringWithMarkup' in value:
                                                markup_list = value['StringWithMarkup']
                                                if markup_list and len(markup_list) > 0:
                                                    smiles = markup_list[0].get('String', 'Not available')
                                                    self.log(f"✓ SMILES found")
                                                    return smiles
                                        break
                                break
                        break
        except Exception as e:
            
            self.log(f"⚠ SMILES: {str(e)}")

        return smiles

    def find_ghs_section(self, sections, path=[]):
        for section in sections:
            heading = section.get('TOCHeading', '')

            if 'GHS Classification' in heading:
                return section

            if 'Section' in section:
                result = self.find_ghs_section(section['Section'], path + [heading])
                if result:
                    return result

        return None

    def fetch_ghs_data(self, cid):
        pictograms = []
        hazard_statements = []

        try:
            url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound/{cid}/JSON"
            response = requests.get(url, timeout=15)

            if response.status_code == 200:
                data = response.json()

                sections = data.get('Record', {}).get('Section', [])
                for section in sections:
                    if section.get('TOCHeading') == 'Safety and Hazards':
                        ghs_section = self.find_ghs_section(section.get('Section', []))

                        if ghs_section:
                            self.log(f"✓ Found GHS section")

                            for info in ghs_section.get('Information', []):
                                info_name = info.get('Name', '')

                                if info_name == 'Pictogram(s)':
                                    value = info.get('Value', {})
                                    string_with_markup = value.get('StringWithMarkup', [])
                                    for item in string_with_markup:
                                        if len(pictograms) >= 3:
                                            break
                                        markup_list = item.get('Markup', [])
                                        for markup in markup_list:
                                            if len(pictograms) >= 3:
                                                break
                                            if markup.get('Type') == 'Icon':
                                                pic_url = markup.get('URL', '')
                                                pic_label = markup.get('Extra', '')
                                                if pic_url:
                                                    pictograms.append({
                                                        'url': pic_url,
                                                        'label': pic_label
                                                    })

                                elif 'GHS Hazard Statement' in info_name or info_name == 'Hazard Statement(s)':
                                    value = info.get('Value', {})

                                    if 'StringValueList' in value:
                                        for statement in value['StringValueList']:
                                            if len(hazard_statements) >= 5:
                                                break
                                            hazard_statements.append(statement)

                                    elif 'StringValue' in value:
                                        if len(hazard_statements) < 5:
                                            hazard_statements.append(value['StringValue'])

                                    elif 'StringWithMarkup' in value:
                                        for item in value['StringWithMarkup']:
                                            if len(hazard_statements) >= 5:
                                                break
                                            if 'String' in item:
                                                hazard_statements.append(item['String'])
                            break

        except Exception as e:
            
            pass

        return pictograms, hazard_statements

    def load_ghs_images(self, pictograms):
        images = []
        labels = []

        for pic in pictograms:
            try:
                url = pic['url']
                label = pic['label']

                gif_url = url.replace('.svg', '.gif')
                response = requests.get(gif_url, timeout=5)

                if response.status_code == 200:
                    img = Image.open(BytesIO(response.content))
                    img = img.resize((100, 100), Image.Resampling.LANCZOS)
                    photo = ImageTk.PhotoImage(img)
                    images.append(photo)
                    labels.append(label)
            except:
                pass

        return images, labels

    def display_ghs_images(self, images, labels):
        for widget in self.hazard_frame.winfo_children():
            widget.destroy()

        if not images:
            self.hazard_label = tk.Label(self.hazard_frame, text="No GHS pictograms",
                                        bg="black", fg="gray")
            self.hazard_label.pack(expand=True)
            return

        img_container = tk.Frame(self.hazard_frame, bg="black")
        img_container.pack(expand=True)

        for idx, (photo, label) in enumerate(zip(images, labels)):
            pic_frame = tk.Frame(img_container, bg="black")
            pic_frame.pack(side="left", padx=8, pady=5)

            img_label = tk.Label(pic_frame, image=photo, bg="black")
            img_label.image = photo
            img_label.pack()

            text_label = tk.Label(pic_frame, text=label, bg="black", fg="white",
                                 font=("Arial", 9, "bold"), wraplength=100)
            text_label.pack()

    def create_excel_file(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )

        if file_path:
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Chemicals"

            headers = ['Sl. No', 'Chemical Name']
            if self.include_cas.get():
                headers.append('CAS No.')
            if self.include_formula.get():
                headers.append('Molecular Formula')
            if self.include_molweight.get():
                headers.append('Molecular Weight')
                headers.append('SI.Unit')
            if self.include_density.get():
                headers.append('Density')
                headers.append('SI.Unit')
            if self.include_quantity.get():
                headers.append('Quantity')
                headers.append('SI.Unit')
            if self.include_equivalence.get():
                headers.append('Equivalence')
            if self.include_iupac.get():
                headers.append('IUPAC Name')
            if self.include_smiles.get():
                headers.append('SMILES')
            if self.include_image_link.get():
                headers.append('Image Link')


            for col, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            sheet.column_dimensions['A'].width = 10
            sheet.column_dimensions['B'].width = 30
            col_idx = 3

            if self.include_cas.get():
                sheet.column_dimensions[get_column_letter(col_idx)].width = 16
                col_idx += 1

            if self.include_formula.get():
                sheet.column_dimensions[get_column_letter(col_idx)].width = 20
                col_idx += 1

            if self.include_molweight.get():
                sheet.column_dimensions[get_column_letter(col_idx)].width = 14  # MW
                col_idx += 1
                sheet.column_dimensions[get_column_letter(col_idx)].width = 12  # MW SI
                col_idx += 1

            if self.include_density.get():
                sheet.column_dimensions[get_column_letter(col_idx)].width = 14  # Density
                col_idx += 1
                sheet.column_dimensions[get_column_letter(col_idx)].width = 16  # Density SI
                col_idx += 1

            if self.include_quantity.get():
                sheet.column_dimensions[get_column_letter(col_idx)].width = 12  # Quantity
                col_idx += 1
                sheet.column_dimensions[get_column_letter(col_idx)].width = 14  # Quantity SI
                col_idx += 1

            if self.include_equivalence.get():
                sheet.column_dimensions[get_column_letter(col_idx)].width = 14
                col_idx += 1

            if self.include_iupac.get():
                sheet.column_dimensions[get_column_letter(col_idx)].width = 45
                col_idx += 1

            if self.include_smiles.get():
                sheet.column_dimensions[get_column_letter(col_idx)].width = 40
                col_idx += 1

            if self.include_image_link.get():
                sheet.column_dimensions[get_column_letter(col_idx)].width = 45
                col_idx += 1


            wb.save(file_path)
            self.excel_file = file_path
            self.file_label.config(text=os.path.basename(file_path), fg="green")
            self.log(f"✓ Excel created: {os.path.basename(file_path)}")
            messagebox.showinfo("Success", f"Excel created: {os.path.basename(file_path)}")

    def load_excel_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )

        if file_path:
            self.excel_file = file_path
            self.file_label.config(text=os.path.basename(file_path), fg="green")
            self.log(f"✓ Excel loaded: {os.path.basename(file_path)}")
            messagebox.showinfo("Success", f"Excel loaded: {os.path.basename(file_path)}")
            wb = load_workbook(file_path)
            
            sheet = wb.active

            headers = [cell.value for cell in sheet[1] if cell.value]

            self.include_cas.set('CAS No.' in headers)
            self.include_formula.set('Molecular Formula' in headers)
            self.include_molweight.set('Molecular Weight' in headers)
            self.include_iupac.set('IUPAC Name' in headers)
            self.include_smiles.set('SMILES' in headers)
            self.include_density.set('Density' in headers)
    
    def fetch_preferred_name(self, cid):
        try:
            url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound/{cid}/JSON"
            response = requests.get(url, timeout=10)

            if response.status_code == 200:
                data = response.json()
                record = data.get("Record", {})
                return record.get("RecordTitle", "Not available")

        except:
            pass

        return None, None
    
    def open_pubchem_page(self):
        if not self.current_data:
            messagebox.showwarning(
                "No compound",
                "Please search for a chemical first."
            )
            return

        cid = self.current_data.get("cid")
        if cid:
            webbrowser.open_new(f"https://pubchem.ncbi.nlm.nih.gov/compound/{cid}")
    
    def is_online(self):
        try:
            requests.get("https://pubchem.ncbi.nlm.nih.gov", timeout=2)
            return True
        except:
            return False

    def normalize_key(self, name: str) -> str:
        return re.sub(r"\s+", " ", name.strip().lower())

    def compute_hash(self, raw_bytes: bytes) -> str:
        return hashlib.sha256(raw_bytes).hexdigest()

    def search_chemical(self):
        raw_query = self.name_entry.get().strip()
        chemical_name = raw_query.lower()
        key = self.normalize_key(raw_query)

        now = time.time()

        if self.search_in_progress:
            self.log("⏳ Search already in progress")
            return

        if now - self.last_search_time < 1:
            self.log("⏳ Please wait before searching again")
            return
        if self.last_searched_query == chemical_name and self.current_data:
            self.log("ℹ Same compound already loaded")
            return
        if not chemical_name:
            messagebox.showwarning("Input Error", "Please enter a chemical name")
            return

        self.clear_results()   # 🔑 ALWAYS reset UI
        self.hide_suggestions()
        self.search_in_progress = True
        self.last_search_time = now

        online = self.is_online()

        cache_key = None

        if key in self.cache:
            cache_key = key
        elif chemical_name in self.cas_index:
            cache_key = self.cas_index[chemical_name]
        elif key in self.iupac_index:
            cache_key = self.iupac_index[key]
        elif raw_query in self.smiles_index:
            cache_key = self.smiles_index[raw_query]

        if cache_key is not None and cache_key in self.cache and not online:
            data = self.cache[cache_key]

            self.log("✓ Loaded from local cache")

            self.current_data = data
            self.last_searched_query = chemical_name

            # Populate UI (same as before)
            self.title_var.set(data["name"])
            self.cas_var.set(data["cas"])
            self.formula_var.set(data["formula"])
            self.molweight_var.set(f'{data["mw"]} {data["mw_u"]}')
            self.density_var.set(
                f'{data["dens"]} {data["dens_u"]}' if data["dens"] is not None else "Not available"
            )

            self.set_text_readonly(self.iupac_text, data["iupac"])
            self.set_text_readonly(self.smiles_text, data["smiles"])

            self.hazard_text.delete(1.0, tk.END)
            if data.get("ghs"):
                for i, stmt in enumerate(data["ghs"], 1):
                    self.hazard_text.insert(tk.END, f"{i}. {stmt}\n")
            else:
                self.hazard_text.insert(tk.END, "No hazard data (cached)")

            try:
                img = Image.open(BytesIO(requests.get(data["img"], timeout=5).content))
                img.thumbnail((500, 320))
                photo = ImageTk.PhotoImage(img)
                self.image_label.config(image=photo, text="")
                self.image_label.image = photo
            except:
                self.image_label.config(text="Offline (no image)", image="")

            self.search_in_progress = False
            return

        now = time.time()

        # Update search state
        self.search_in_progress = True
        self.last_search_time = now

        self.suggestion_confirmed = False

        self.log(f"\n{'='*40}")
        self.log(f"Searching: {chemical_name}")
        self.log(f"{'='*40}")

        try:
            search_url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/{chemical_name}/JSON"
            response = requests.get(search_url, timeout=10)

            if response.status_code != 200:
                self.log(f"✗ Chemical not found")
                messagebox.showerror("Not Found", f"'{chemical_name}' not found")
                return

            data = response.json()
            cid = data['PC_Compounds'][0]['id']['id']['cid']
            self.log(f"✓ CID: {cid}")
            preferred_name = self.fetch_preferred_name(cid)
            self.title_var.set(preferred_name)


            molecular_weight_value, molecular_weight_unit = self.fetch_molecular_weight(cid)
            density_value, density_unit = self.fetch_density(cid)
            if density_value is not None:
                self.density_var.set(f"{density_value} {density_unit}")
                self.log(f"✓ Density: {density_value} {density_unit}")
            else:
                self.density_var.set("Not available")
                self.log(f"none")

            molecular_formula = "Not available"

            try:
                props = data['PC_Compounds'][0]['props']
                for prop in props:
                    if prop['urn']['label'] == 'Molecular Formula':
                        molecular_formula = prop['value'].get('sval', "Not available")
                        break
            except:
                pass

            self.formula_var.set(molecular_formula)
            self.molweight_var.set(f"{molecular_weight_value} {molecular_weight_unit}")

            self.log(f"✓ Formula: {molecular_formula}")
            self.log(f"✓ Mol.Weight: {molecular_weight_value} {molecular_weight_unit}")

            # Get CAS number
            cas_number = "Not available"
            try:
                syn_url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{cid}/synonyms/JSON"
                syn_response = requests.get(syn_url, timeout=10)
                if syn_response.status_code == 200:
                    syn_data = syn_response.json()
                    synonyms = syn_data['InformationList']['Information'][0]['Synonym']
                    for syn in synonyms:
                        if '-' in syn and syn.replace('-', '').isdigit():
                            parts = syn.split('-')
                            if len(parts) == 3 and parts[2].isdigit() and len(parts[2]) == 1:
                                cas_number = syn
                                break
            except:
                pass

            self.cas_var.set(cas_number)
            self.log(f"✓ CAS: {cas_number}")

            iupac_name = self.fetch_iupac_name(cid)
            smiles = self.fetch_smiles(cid)
            self.set_text_readonly(self.iupac_text, iupac_name)
            self.set_text_readonly(self.smiles_text, smiles)

            # Get structure image
            image_url = f"https://pubchem.ncbi.nlm.nih.gov/image/imgsrv.fcgi?cid={cid}&t=l"
            try:
                img_response = requests.get(image_url, timeout=10)
                if img_response.status_code == 200:
                    img = Image.open(BytesIO(img_response.content))
                    img.thumbnail((500, 320))
                    photo = ImageTk.PhotoImage(img)
                    self.image_label.config(image=photo, text="")
                    self.image_label.image = photo
                    self.log(f"✓ Image loaded")
            except:
                pass

            # Fetch GHS data
            pictograms, hazard_statements = self.fetch_ghs_data(cid)

            if pictograms:
                images, labels = self.load_ghs_images(pictograms)
                if images:
                    self.display_ghs_images(images, labels)
                    self.log(f"✓ GHS: {len(images)} pictogram(s)")

            if hazard_statements:
                self.hazard_text.delete(1.0, tk.END)
                for idx, statement in enumerate(hazard_statements[:5], 1):
                    self.hazard_text.insert(tk.END, f"{idx}. {statement}\n")
                self.log(f"✓ Hazards: {len(hazard_statements[:5])}")
            else:
                self.hazard_text.delete(1.0, tk.END)
                self.hazard_text.insert(tk.END, "No hazards available")

            # Store current data
            self.current_data = {
                'name': preferred_name,
                'cid': cid,   
                'cas': cas_number,
                'formula': molecular_formula,
                'molweight_value': molecular_weight_value,  
                'molweight_unit': 'g/mol',
                'density_value': density_value,               
                'density_unit': density_unit,      
                'iupac': iupac_name,
                'smiles': smiles,
                'image': image_url
            }

            # ---------- SAVE TO LOCAL CACHE ----------
            key = self.normalize_key(preferred_name)

            if key not in self.cache:
                self.cache[key] = {
                    "cid": cid,
                    "name": preferred_name,
                    "cas": cas_number,
                    "formula": molecular_formula,
                    "mw": molecular_weight_value,
                    "mw_u": "g/mol",
                    "dens": density_value,
                    "dens_u": density_unit,
                    "iupac": iupac_name,
                    "smiles": smiles,
                    "ghs": hazard_statements[:2] if hazard_statements else [],
                    "img": image_url,
                    "ts": int(time.time())
                }

                try:
                    raw = json.dumps(
                        self.cache,
                        separators=(",", ":"),
                        ensure_ascii=False
                    ).encode("utf-8")

                    with open(CACHE_FILE, "wb") as f:
                        f.write(raw)

                    with open(CACHE_SIG_FILE, "w") as sig:
                        sig.write(self.compute_hash(raw))

                    self.log("✓ Cached locally")

                except:
                    self.log("⚠ Failed to save cache")

            self.last_searched_query = chemical_name

            self.log(f"{'='*40}")
            self.log(f"✓ Ready to save!")
            self.log(f"{'='*40}\n")
        

        except Exception as e:
            self.search_in_progress = False
            self.log_error("Error", str(e), f"Type: {type(e).__name__}")
            messagebox.showerror("Error", f"Error: {str(e)}")
        finally:
            self.search_in_progress = False

    def silent_refresh(self, key):
        try:
            # Quick online test
            requests.get("https://pubchem.ncbi.nlm.nih.gov", timeout=2)

            data = self.cache[key]
            cid = data["cid"]
            updated = False

            if not data.get("smiles"):
                data["smiles"] = self.fetch_smiles(cid)
                updated = True

            if not data.get("ghs"):
                _, hazards = self.fetch_ghs_data(cid)
                data["ghs"] = hazards[:2]
                updated = True

            if updated:
                raw = json.dumps(
                    self.cache,
                    separators=(",", ":"),
                    ensure_ascii=False
                ).encode("utf-8")

                with open(CACHE_FILE, "wb") as f:
                    f.write(raw)

                with open(CACHE_SIG_FILE, "w") as sig:
                    sig.write(self.compute_hash(raw))

        except:
            pass

    
    def set_text_readonly(self, widget, value):
        widget.config(state="normal")
        widget.delete("1.0", tk.END)
        widget.insert("1.0", value)
        widget.config(state="disabled")

    def fetch_molecular_weight(self, cid):
        try:
            url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{cid}/property/MolecularWeight/JSON"
            response = requests.get(url, timeout=10)

            if response.status_code == 200:
                data = response.json()
                mw_raw = data['PropertyTable']['Properties'][0]['MolecularWeight']
                mw = float(mw_raw)
                return mw, "g/mol"


        except Exception as e:
            
            self.log(f"⚠ MWT error: {str(e)}")

        return None, None

    def add_to_excel(self):
        if not self.excel_file:
            messagebox.showwarning("No File", "Create or load Excel first")
            return

        if not self.current_data:
            messagebox.showwarning("No Data", "Search for a chemical first")
            return

        try:
            wb = load_workbook(self.excel_file)
            sheet = wb.active

            next_row = sheet.max_row + 1
            sl_no = next_row - 1

            col = 1
            sheet.cell(row=next_row, column=col).value = sl_no
            col += 1

            sheet.cell(row=next_row, column=col).value = self.current_data['name']
            col += 1

            if self.include_cas.get():
                sheet.cell(row=next_row, column=col).value = self.current_data['cas']
                col += 1

            if self.include_formula.get():
                sheet.cell(row=next_row, column=col).value = self.current_data['formula']
                col += 1

            if self.include_molweight.get():
                sheet.cell(row=next_row, column=col).value = self.current_data['molweight_value']
                col += 1
                sheet.cell(row=next_row, column=col).value = self.current_data['molweight_unit']
                col += 1
            
            if self.include_density.get():
                sheet.cell(row=next_row, column=col).value = self.current_data['density_value']
                col += 1
                sheet.cell(row=next_row, column=col).value = self.current_data['density_unit']
                col += 1

            if self.include_quantity.get():
                  col += 2

            if self.include_equivalence.get():
                col += 1  # leave cell empty intentionally

            if self.include_iupac.get():
                sheet.cell(row=next_row, column=col).value = self.current_data['iupac']
                col += 1

            if self.include_smiles.get():
                sheet.cell(row=next_row, column=col).value = self.current_data['smiles']
                col += 1    

            if self.include_image_link.get():
                sheet.cell(row=next_row, column=col).value = self.current_data['image']
                col += 1

            wb.save(self.excel_file)
            self.log(f"\n✓✓✓ SAVED! ✓✓✓")
            messagebox.showinfo("Success", f"Added '{self.current_data['name']}'!")

            self.name_entry.delete(0, tk.END)
            self.clear_results()

        except PermissionError:
            self.log_error("File Locked", "Close Excel file first", "")
            messagebox.showerror("Locked", "Close the Excel file first")

        except Exception as e:
            
            self.log_error("Save Error", str(e), "")
            messagebox.showerror("Error", f"Save failed: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = PubChemScraperApp(root)
    root.mainloop()
